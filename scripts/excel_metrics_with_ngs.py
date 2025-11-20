#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import json
import os
import re
from typing import Dict, Any, List, Tuple, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# --------- Ρυθμίσεις Ομάδων Κατηγοριών ---------
GROUPS = {
    "mipnerf360_outdoor_scenes": ["bicycle", "flowers", "garden", "stump", "treehill"],
    "mipnerf360_indoor_scenes": ["room", "counter", "kitchen", "bonsai"],
    "tanks_and_temples_scenes": ["truck", "train"],
    "deep_blending_scenes": ["drjohnson", "playroom"],
}

def load_results_json(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def collect_metrics(eval_dir: str) -> pd.DataFrame:
    """
    Επιστρέφει DataFrame (tidy) με στήλες:
    ['category', 'step', 'SSIM', 'PSNR', 'LPIPS', 'number_of_gaussians']
    """
    rows: List[Dict[str, Any]] = []
    for name in sorted(os.listdir(eval_dir)):
        cat_dir = os.path.join(eval_dir, name)
        if not os.path.isdir(cat_dir):
            continue

        results_path = os.path.join(cat_dir, "results.json")
        if not os.path.isfile(results_path):
            continue

        try:
            data = load_results_json(results_path)
        except Exception as e:
            print(f"[warn] Δεν διαβάστηκε το {results_path}: {e}")
            continue

        # Retrieve the number of vertices from the PLY file
        ply_dir = os.path.join(cat_dir, "point_cloud", "iteration_30000")
        ply_file = Path(ply_dir) / "point_cloud.ply"
        vertex_count = count_vertices_in_ply(ply_file)

        # Ensure vertex_count is an integer, set to 0 if not found
        vertex_count = int(vertex_count) if vertex_count is not None else 0

        for key, metrics in data.items():
            m = re.match(r"ours_(\d+)", key)
            if not m:
                continue
            step = int(m.group(1))
            if step not in (7000, 30000):
                continue
            rows.append(
                {
                    "category": name,
                    "step": step,
                    "SSIM": metrics.get("SSIM"),
                    "PSNR": metrics.get("PSNR"),
                    "LPIPS": metrics.get("LPIPS"),
                    "number_of_gaussians": vertex_count,  # Now always an integer
                }
            )

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["category", "step"]).reset_index(drop=True)
    return df

def count_vertices_in_ply(ply_path: Path) -> Optional[int]:
    """Parse PLY header to get number of vertices. Returns None if not found or file invalid."""
    try:
        with ply_path.open('rb') as f:
            vertex_count = None
            while True:
                line = f.readline()
                if not line:
                    break
                try:
                    s = line.decode('utf-8', errors='ignore').strip()
                except Exception:
                    s = ''
                if s.startswith('element vertex'):
                    parts = s.split()
                    if len(parts) >= 3 and parts[0] == 'element' and parts[1] == 'vertex':
                        try:
                            vertex_count = int(parts[2])
                        except ValueError:
                            pass
                if s == 'end_header':
                    break
            return vertex_count
    except Exception as e:
        return None

def to_wide(df: pd.DataFrame) -> pd.DataFrame:
    """Wide μορφή: μία γραμμή ανά category, στήλες SSIM_7000, SSIM_30000, κ.λπ."""
    if df.empty:
        return df
    wide = (
        df.pivot_table(
            index="category",
            columns="step",
            values=["SSIM", "PSNR", "LPIPS", "number_of_gaussians"],
        )
        .sort_index(axis=1)
        .reset_index()
    )
    wide.columns = [
        "category" if c[0] == "category" else f"{c[0]}_{c[1]}"
        for c in wide.columns
    ]

    # Ταξινόμηση κατηγοριών με βάση τις ομάδες
    ordered = []
    present = set(wide["category"].tolist())
    for _, cats in GROUPS.items():
        ordered.extend([c for c in cats if c in present])
    remaining = sorted([c for c in present if c not in ordered])
    order_index = {c: i for i, c in enumerate(ordered + remaining)}

    wide = wide.sort_values(
        by="category",
        key=lambda s: s.map(order_index)
    ).reset_index(drop=True)

    return wide

# --------- Excel Formatting ---------

def autofit_columns(ws):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            txt = "" if v is None else str(v)
            widths[i] = max(widths.get(i, 0), len(txt))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 40)

def write_formatted_excel(wide: pd.DataFrame, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics"

    # Header
    headers = list(wide.columns)
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="BBBBBB")

    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Γράψιμο δεδομένων κατά ομάδες με κενή γραμμή ανάμεσα
    row_idx = 2

    # Χάρτης: category -> σειρά wide
    by_cat = {row["category"]: row for _, row in wide.iterrows()}

    def write_row(row_values: List[Any], r: int):
        for j, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            if j == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = right
                if isinstance(val, int):  # Ensure it's formatted as an integer
                    cell.number_format = '0'  # Format as integer without decimals
                elif isinstance(val, float):
                    if "LPIPS" in ws.cell(row=1, column=j).value:  # SSIM / LPIPS specific formatting
                        cell.number_format = "0.0000"
                    else:
                        cell.number_format = "0.00"  # PSNR

    # Σειρά ομάδων
    for group_name, cats in GROUPS.items():
        group_rows = [by_cat[c] for c in cats if c in by_cat]
        if not group_rows:
            continue
        for r in group_rows:
            write_row([r.get(h) for h in headers], row_idx)
            row_idx += 1
        row_idx += 1

    # Ό,τι περίσσεψε
    already = set(sum(GROUPS.values(), []))
    leftover = [c for c in wide["category"].tolist() if c not in already]
    for c in leftover:
        r = by_cat[c]
        write_row([r.get(h) for h in headers], row_idx)
        row_idx += 1

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto width
    autofit_columns(ws)

    # Save
    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description="Συλλογή μετρικών από eval/*/results.json σε Excel με μορφοποίηση.")
    parser.add_argument(
        "--eval_dir",
        default="eval",
        help="Διαδρομή προς τον φάκελο eval (default: ./eval)",
    )
    parser.add_argument(
        "--filename",
        default="eval_metrics.xlsx",
        help="Όνομα του Excel αρχείου (default: eval_metrics.xlsx). Αποθηκεύεται ΜΕΣΑ στο eval_dir.",
    )
    args = parser.parse_args()

    eval_dir = os.path.abspath(args.eval_dir)
    if not os.path.isdir(eval_dir):
        raise FileNotFoundError(f"Δεν βρέθηκε ο φάκελος: {eval_dir}")

    tidy = collect_metrics(eval_dir)
    if tidy.empty:
        print("Δεν βρέθηκαν μετρικές (ours_7000 / ours_30000).")
        return

    wide = to_wide(tidy)

    out_path = os.path.join(eval_dir, args.filename)
    write_formatted_excel(wide, out_path)

    print(f"OK! Το Excel αποθηκεύτηκε εδώ: {out_path}")


if __name__ == "__main__":
    main()
